import streamlit as st
import google.genai as genai
import requests
import csv
import json
import re
import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="EXD Keyword Research Tool", page_icon="⚡", layout="centered")

# ── Password protection ───────────────────────────────────────────
def check_password():
    try:
        app_password = st.secrets.get("APP_PASSWORD", "") or os.environ.get("APP_PASSWORD", "")
    except:
        app_password = os.environ.get("APP_PASSWORD", "")

    if not app_password:
        return True  # No password set — allow access

    if st.session_state.get("authenticated"):
        return True

    st.markdown("""
    <style>
    .login-box { max-width: 380px; margin: 6rem auto; text-align: center; }
    .login-box h2 { font-size: 22px; font-weight: 700; color: #fff; margin-bottom: 0.5rem; }
    .login-box p { font-size: 14px; color: #888; margin-bottom: 2rem; }
    </style>
    <div class="login-box">
        <div style="font-size:36px;margin-bottom:1rem">⚡</div>
        <h2>EXD Keyword Research Tool</h2>
        <p>Enter your password to continue</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        pwd = st.text_input("Password", type="password", label_visibility="collapsed",
                            placeholder="Enter password...")
        if st.button("Login", use_container_width=True):
            if pwd == app_password:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password. Please try again.")
    return False

if not check_password():
    st.stop()
# ─────────────────────────────────────────────────────────────────

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
  html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
  .block-container { max-width: 800px; padding: 2rem 2rem 4rem; }
  .section-hdr { display:flex; align-items:center; gap:12px; margin:2.5rem 0 1rem; }
  .section-num {
    min-width:28px; height:28px; border-radius:50%;
    background:#2a2a2a; color:#fff; font-size:13px; font-weight:600;
    display:flex; align-items:center; justify-content:center;
  }
  .section-ttl { font-size:15px; font-weight:600; color:#fff; }
  .section-sub { font-size:13px; color:#666; font-weight:400; }
  div[data-testid="stButton"] > button {
    background:#1e1e1e !important; border:1.5px solid #333 !important;
    border-radius:10px !important; color:#fff !important;
    font-weight:500 !important; transition:border-color 0.15s !important;
  }
  div[data-testid="stButton"] > button:hover {
    border-color:#f97316 !important; background:#1e1e1e !important; color:#fff !important;
  }
  .gen-btn div[data-testid="stButton"] > button {
    background:#f97316 !important; border:none !important;
    border-radius:10px !important; font-size:15px !important;
    font-weight:600 !important; color:#fff !important; padding:0.75rem !important;
  }
  .gen-btn div[data-testid="stButton"] > button:hover { background:#ea6c0a !important; }
  .add-btn div[data-testid="stButton"] > button {
    background:#f97316 !important; border:none !important;
    border-radius:8px !important; color:#fff !important; font-weight:600 !important;
  }
  .add-btn div[data-testid="stButton"] > button:hover { background:#ea6c0a !important; }
  div[data-testid="stDownloadButton"] > button {
    background:#f97316 !important; color:#fff !important;
    border:none !important; border-radius:10px !important; font-weight:600 !important;
  }
  div[data-testid="stDownloadButton"] > button:hover { background:#ea6c0a !important; }
  .tag-pill {
    display:inline-flex; align-items:center; gap:6px;
    background:#1e1e1e; border:1px solid #444; border-radius:20px;
    padding:5px 14px; font-size:13px; color:#ddd; margin:3px;
    cursor:pointer;
  }
  hr { border-color:#2a2a2a !important; }
  #MainMenu, footer { visibility:hidden; }
</style>
""", unsafe_allow_html=True)

DFS_LOCATIONS = {
    "Saudi Arabia": "Saudi Arabia",
    "UAE":          "United Arab Emirates",
    "Egypt":        "Egypt",
    "Kuwait":       "Kuwait",
    "Qatar":        "Qatar",
    "Bahrain":      "Bahrain",
    "Oman":         "Oman",
    "Jordan":       "Jordan",
    "Lebanon":      "Lebanon",
    "United Kingdom": "United Kingdom",
    "United States":  "United States",
    "France":       "France",
    "Germany":      "Germany",
    "India":        "India",
    "China":        "China",
    "Japan":        "Japan",
}
MARKETS = {
    "Saudi Arabia":"sa","UAE":"ae","Egypt":"eg","Kuwait":"kw","Qatar":"qa",
    "Bahrain":"bh","Oman":"om","Jordan":"jo","Lebanon":"lb",
    "United Kingdom":"uk","United States":"us","France":"fr",
    "Germany":"de","India":"in","China":"cn","Japan":"jp"
}
LANGUAGES = [
    {"code":"en","name":"English", "native":"English",  "flag":"🇬🇧"},
    {"code":"ar","name":"Arabic",  "native":"العربية",   "flag":"🇸🇦"},
    {"code":"fr","name":"French",  "native":"Français",  "flag":"🇫🇷"},
    {"code":"de","name":"German",  "native":"Deutsch",   "flag":"🇩🇪"},
    {"code":"hi","name":"Hindi",   "native":"हिन्दी",    "flag":"🇮🇳"},
    {"code":"zh","name":"Chinese", "native":"简体中文",  "flag":"🇨🇳"},
    {"code":"ja","name":"Japanese","native":"日本語",    "flag":"🇯🇵"},
]
QTYPES = [
    {"code":"direct",   "title":"Direct",   "desc":"Short, high-intent. 1–3 words."},
    {"code":"longtail", "title":"Long-tail", "desc":"Specific phrases. 4–8 words."},
    {"code":"mix",      "title":"Mix",       "desc":"Balanced blend of both."},
]
QTYPE_INSTRUCTIONS = {
    "direct":   "Generate SHORT HIGH-INTENT keywords only. Each must be 1-3 words maximum.",
    "longtail": "Generate LONG-TAIL keywords only. Each must be 4-8 words. Focus on specific questions and use-cases.",
    "mix":      "Generate a MIX: roughly half 1-3 words (direct) and half 4-8 words (long-tail).",
}

if "nav_tags"       not in st.session_state: st.session_state.nav_tags = []
if "selected_langs" not in st.session_state: st.session_state.selected_langs = ["en"]
if "selected_qt"    not in st.session_state: st.session_state.selected_qt = "mix"

try:
    gemini_key   = st.secrets.get("GEMINI_KEY", "")   or os.environ.get("GEMINI_KEY", "")
    dfs_login    = st.secrets.get("DFS_LOGIN", "")    or os.environ.get("DFS_LOGIN", "")
    dfs_password = st.secrets.get("DFS_PASSWORD", "") or os.environ.get("DFS_PASSWORD", "")
except:
    gemini_key   = os.environ.get("GEMINI_KEY", "")
    dfs_login    = os.environ.get("DFS_LOGIN", "")
    dfs_password = os.environ.get("DFS_PASSWORD", "")

with st.sidebar:
    st.markdown("## ⚡ EXD Keywords")
    st.markdown("Generate pitch keyword lists in seconds.")
    st.divider()
    st.markdown("**How to use**")
    st.caption("1. Fill in account setup")
    st.caption("2. Select languages")
    st.caption("3. Choose keyword type")
    st.caption("4. Add tags")
    st.caption("5. Click Generate")
    st.caption("6. Download Excel or CSV")
    st.divider()
    st.markdown("**API Keys**")
    if gemini_key:
        st.success("Gemini ✓")
    else:
        gemini_key = st.text_input("Gemini API key", type="password")
    st.caption("DataForSEO credentials (optional — for search volumes)")
    if dfs_login and dfs_password:
        st.success("DataForSEO ✓")
        # Allow override in case secrets are wrong
        override = st.checkbox("Override credentials")
        if override:
            dfs_login    = st.text_input("DataForSEO login (email)", value=dfs_login, key="dfs_login_input")
            dfs_password = st.text_input("DataForSEO password", type="password", key="dfs_pass_input")
    else:
        dfs_login    = st.text_input("DataForSEO login (email)", key="dfs_login_input")
        dfs_password = st.text_input("DataForSEO password", type="password", key="dfs_pass_input")

st.markdown("# ⚡ EXD Keyword Research Tool")
st.markdown("Generate, validate, and export pitch keyword lists in seconds.")
st.divider()

# SECTION 1
st.markdown('<div class="section-hdr"><div class="section-num">1</div><div class="section-ttl">Account Setup</div></div>', unsafe_allow_html=True)
c1, c2 = st.columns(2)
with c1:
    client_name  = st.text_input("Client name", placeholder="e.g. Al Shifa Honey")
with c2:
    market_label = st.selectbox("Target market", list(MARKETS.keys()), index=1)
website = st.text_input("Client website", placeholder="https://www.alshifa.com  (optional — validates keyword relevance)")

# ── Mode switcher ─────────────────────────────────────────────────
st.divider()
if "mode" not in st.session_state:
    st.session_state.mode = "keywords"

m1, m2 = st.columns(2)
with m1:
    kw_active  = st.session_state.mode == "keywords"
    kw_border  = "#f97316" if kw_active else "#333"
    kw_tick    = " ✓" if kw_active else ""
    st.markdown(f"""<div style="background:#1e1e1e;border:2px solid {kw_border};border-radius:10px 10px 0 0;
        padding:14px;text-align:center;">
        <div style="font-size:20px">🔍</div>
        <div style="font-size:13px;font-weight:600;color:#fff;margin-top:4px">Generate Keywords<span style="color:#f97316">{kw_tick}</span></div>
        <div style="font-size:11px;color:#888;margin-top:3px">Branded & generic search keywords</div>
    </div>""", unsafe_allow_html=True)
    kw_bg = "#f97316" if kw_active else "#2a2a2a"
    st.markdown(f"""<style>div[data-testid="column"]:nth-of-type(1) div[data-testid="stButton"] > button {{
        background:{kw_bg} !important; border:2px solid {kw_border} !important;
        border-top:none !important; border-radius:0 0 10px 10px !important;
        color:#fff !important; font-size:12px !important; padding:5px !important; width:100% !important;
    }}</style>""", unsafe_allow_html=True)
    if st.button("Select", key="mode_kw"):
        st.session_state.mode = "keywords"
        st.rerun()

with m2:
    pr_active  = st.session_state.mode == "prompts"
    pr_border  = "#f97316" if pr_active else "#333"
    pr_tick    = " ✓" if pr_active else ""
    st.markdown(f"""<div style="background:#1e1e1e;border:2px solid {pr_border};border-radius:10px 10px 0 0;
        padding:14px;text-align:center;">
        <div style="font-size:20px">🤖</div>
        <div style="font-size:13px;font-weight:600;color:#fff;margin-top:4px">Generate Prompts<span style="color:#f97316">{pr_tick}</span></div>
        <div style="font-size:11px;color:#888;margin-top:3px">AI search prompts for ChatGPT, Gemini & more</div>
    </div>""", unsafe_allow_html=True)
    pr_bg = "#f97316" if pr_active else "#2a2a2a"
    st.markdown(f"""<style>div[data-testid="column"]:nth-of-type(2) div[data-testid="stButton"] > button {{
        background:{pr_bg} !important; border:2px solid {pr_border} !important;
        border-top:none !important; border-radius:0 0 10px 10px !important;
        color:#fff !important; font-size:12px !important; padding:5px !important; width:100% !important;
    }}</style>""", unsafe_allow_html=True)
    if st.button("Select", key="mode_pr"):
        st.session_state.mode = "prompts"
        st.rerun()

st.divider()

# ── Mode-specific fields ──────────────────────────────────────────
if st.session_state.mode == "keywords":
    c3, c4 = st.columns(2)
    with c3:
        branded_count = st.number_input("Branded keywords", min_value=0, max_value=1000, value=10, step=1)
    with c4:
        generic_count = st.number_input("Generic keywords", min_value=0, max_value=1000, value=30, step=1)
else:
    prompt_count = st.number_input("Number of prompts per language", min_value=1, max_value=200, value=20, step=1)

# SECTION 2
st.markdown('<div class="section-hdr"><div class="section-num">2</div><div class="section-ttl">Languages</div></div>', unsafe_allow_html=True)
st.caption("Click to select — multiple languages supported.")
lang_cols = st.columns(len(LANGUAGES))
for i, lang in enumerate(LANGUAGES):
    with lang_cols[i]:
        is_active = lang["code"] in st.session_state.selected_langs
        border = "#f97316" if is_active else "#333"
        tick = "✓" if is_active else ""
        st.markdown(f"""<div style="background:#1e1e1e;border:1.5px solid {border};
            border-radius:10px 10px 0 0;padding:10px 4px 6px;text-align:center;">
            <div style="font-size:20px">{lang['flag']}</div>
            <div style="font-size:11px;font-weight:600;color:#fff;margin-top:3px">{lang['name']}</div>
            <div style="font-size:10px;color:#666">{lang['native']}</div>
            <div style="font-size:11px;color:#f97316;font-weight:700;min-height:14px;margin-top:3px">{tick}</div>
        </div>""", unsafe_allow_html=True)
        btn_bg = "#f97316" if is_active else "#2a2a2a"
        st.markdown(f"""<style>div[data-testid="column"]:nth-of-type({i+1}) div[data-testid="stButton"] > button {{
            background:{btn_bg} !important; border:1.5px solid {border} !important;
            border-top:none !important; border-radius:0 0 10px 10px !important;
            color:#fff !important; font-size:12px !important; font-weight:400 !important;
            padding:4px !important; width:100% !important;
        }}</style>""", unsafe_allow_html=True)
        if st.button("Select", key=f"lang_{lang['code']}"):
            if lang["code"] in st.session_state.selected_langs:
                if len(st.session_state.selected_langs) > 1:
                    st.session_state.selected_langs.remove(lang["code"])
            else:
                st.session_state.selected_langs.append(lang["code"])
            st.rerun()
selected_lang_names = [l["name"] for l in LANGUAGES if l["code"] in st.session_state.selected_langs]
if len(selected_lang_names) > 1:
    st.caption(f"✓ {len(selected_lang_names)} languages: {', '.join(selected_lang_names)}")

# SECTION 3
if st.session_state.mode == "keywords":
    st.markdown('<div class="section-hdr"><div class="section-num">3</div><div class="section-ttl">Keyword Type</div></div>', unsafe_allow_html=True)
    qt_cols = st.columns(3)
    for i, qt in enumerate(QTYPES):
        with qt_cols[i]:
            is_active = st.session_state.selected_qt == qt["code"]
            border = "#f97316" if is_active else "#333"
            tick = " ✓" if is_active else ""
            st.markdown(f"""<div style="background:#1e1e1e;border:1.5px solid {border};
                border-radius:10px 10px 0 0;padding:14px 12px 8px;min-height:75px;">
                <div style="font-size:13px;font-weight:600;color:#fff">{qt['title']}<span style="color:#f97316">{tick}</span></div>
                <div style="font-size:11px;color:#888;margin-top:5px;line-height:1.4">{qt['desc']}</div>
            </div>""", unsafe_allow_html=True)
            btn_bg = "#f97316" if is_active else "#2a2a2a"
            st.markdown(f"""<style>div[data-testid="column"]:nth-of-type({i+1}) div[data-testid="stButton"] > button {{
                background:{btn_bg} !important; border:1.5px solid {border} !important;
                border-top:none !important; border-radius:0 0 10px 10px !important;
                color:#fff !important; font-size:12px !important; font-weight:400 !important;
                padding:4px !important; width:100% !important;
            }}</style>""", unsafe_allow_html=True)
            if st.button("Select", key=f"qt_{qt['code']}"):
                st.session_state.selected_qt = qt["code"]
                st.rerun()
else:
    st.markdown('<div class="section-hdr"><div class="section-num">3</div><div class="section-ttl">Prompt Intent</div></div>', unsafe_allow_html=True)
    INTENTS = [
        {"code":"informational", "title":"Informational", "desc":"What is... / How does... / Tell me about..."},
        {"code":"comparison",    "title":"Comparison",    "desc":"What's the best... vs... / Compare..."},
        {"code":"recommendation","title":"Recommendation","desc":"Suggest a... / What should I use for..."},
        {"code":"action",        "title":"Action",        "desc":"Help me find... / Give me a list of..."},
    ]
    if "selected_intent" not in st.session_state:
        st.session_state.selected_intent = ["informational"]
    intent_cols = st.columns(4)
    for i, intent in enumerate(INTENTS):
        with intent_cols[i]:
            is_active = intent["code"] in st.session_state.selected_intent
            border = "#f97316" if is_active else "#333"
            tick = " ✓" if is_active else ""
            st.markdown(f"""<div style="background:#1e1e1e;border:1.5px solid {border};
                border-radius:10px 10px 0 0;padding:12px 8px 8px;min-height:75px;">
                <div style="font-size:12px;font-weight:600;color:#fff">{intent['title']}<span style="color:#f97316">{tick}</span></div>
                <div style="font-size:10px;color:#888;margin-top:4px;line-height:1.4">{intent['desc']}</div>
            </div>""", unsafe_allow_html=True)
            btn_bg = "#f97316" if is_active else "#2a2a2a"
            st.markdown(f"""<style>div[data-testid="column"]:nth-of-type({i+1}) div[data-testid="stButton"] > button {{
                background:{btn_bg} !important; border:1.5px solid {border} !important;
                border-top:none !important; border-radius:0 0 10px 10px !important;
                color:#fff !important; font-size:11px !important; font-weight:400 !important;
                padding:4px !important; width:100% !important;
            }}</style>""", unsafe_allow_html=True)
            if st.button("Select", key=f"intent_{intent['code']}"):
                if intent["code"] in st.session_state.selected_intent:
                    if len(st.session_state.selected_intent) > 1:
                        st.session_state.selected_intent.remove(intent["code"])
                else:
                    st.session_state.selected_intent.append(intent["code"])
                st.rerun()

# SECTION 4
st.markdown(f'<div class="section-hdr"><div class="section-num">4</div><div class="section-ttl">Tags</div></div>', unsafe_allow_html=True)

st.markdown("""<style>
div[data-testid="stForm"] { border: none !important; padding: 0 !important; }
div[data-testid="stForm"] div[data-testid="stButton"] > button {
    background:#f97316 !important; border:none !important;
    border-radius:8px !important; color:#fff !important;
    font-weight:600 !important; width:100% !important;
}
div[data-testid="stForm"] div[data-testid="stButton"] > button:hover { background:#ea6c0a !important; }
</style>""", unsafe_allow_html=True)

with st.form(key="tag_form", clear_on_submit=True):
    tc, bc = st.columns([5, 1])
    with tc:
        new_tag = st.text_input("tag", placeholder="e.g. Honey, Royal Jelly...", label_visibility="collapsed")
    with bc:
        add_clicked = st.form_submit_button("+ Add")
    if add_clicked and new_tag.strip():
        tag_val = new_tag.strip()
        if tag_val not in st.session_state.nav_tags:
            st.session_state.nav_tags.append(tag_val)
        st.rerun()

if st.session_state.nav_tags:
    pills = "".join(
        f'<span style="display:inline-flex;align-items:center;gap:6px;background:#1e1e1e;'
        f'border:1px solid #444;border-radius:20px;padding:5px 14px;font-size:13px;'
        f'color:#ddd;margin:3px 3px;">{tag}</span>'
        for tag in st.session_state.nav_tags
    )
    st.markdown(f'<div style="display:flex;flex-wrap:wrap;margin-top:6px;">{pills}</div>', unsafe_allow_html=True)
    st.markdown("<p style='font-size:12px;color:#666;margin:10px 0 4px;'>Click to remove:</p>", unsafe_allow_html=True)
    st.markdown("""<style>
    .rm-tag-row div[data-testid="stButton"] > button {
        background:transparent !important; border:1px solid #444 !important;
        border-radius:20px !important; color:#888 !important;
        font-size:12px !important; padding:3px 12px !important;
        height:auto !important; width:auto !important; margin:2px !important;
    }
    .rm-tag-row div[data-testid="stButton"] > button:hover {
        border-color:#f97316 !important; color:#f97316 !important;
    }
    </style>""", unsafe_allow_html=True)
    st.markdown('<div class="rm-tag-row">', unsafe_allow_html=True)
    for idx, tag in enumerate(st.session_state.nav_tags):
        col, _ = st.columns([1.2, 8 - min(len(st.session_state.nav_tags) * 1.2, 7)])
        with col:
            if st.button(f"✕ {tag}", key=f"rm_{idx}"):
                st.session_state.nav_tags.pop(idx)
                st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
else:
    st.caption("No tags yet — type above and press Enter or click + Add")

# SECTION 5
seed_label       = "Seed Keywords" if st.session_state.mode == "keywords" else "Seed Prompts"
seed_placeholder = "One per line — specific keywords that must be included" if st.session_state.mode == "keywords" else "One per line — specific prompts that must be included"
st.markdown(f'<div class="section-hdr"><div class="section-num">5</div><div class="section-ttl">{seed_label} <span class="section-sub">— optional</span></div></div>', unsafe_allow_html=True)
seeds_input = st.text_area("seeds", height=100, label_visibility="collapsed", placeholder=seed_placeholder)

st.divider()

if "generating" not in st.session_state:
    st.session_state.generating = False

st.markdown("""<style>
.gen-btn div[data-testid="stButton"] > button {
    background:#f97316 !important; border:none !important;
    border-radius:10px !important; font-size:15px !important;
    font-weight:600 !important; color:#fff !important; padding:0.7rem !important;
}
.gen-btn div[data-testid="stButton"] > button:hover { background:#ea6c0a !important; }
</style>""", unsafe_allow_html=True)

st.markdown('<div class="gen-btn">', unsafe_allow_html=True)
btn_label = "⚡  Generate Keywords" if st.session_state.mode == "keywords" else "⚡  Generate Prompts"
generate = st.button(btn_label, key="generate_btn", use_container_width=True)
st.markdown('</div>', unsafe_allow_html=True)

if generate:
    st.session_state.generating = True
    st.session_state.stop_requested = False
    st.session_state.results = None

    if not gemini_key:
        st.session_state.generating = False
        st.error("Please enter your Gemini API key in the sidebar.")
        st.stop()
    if not client_name:
        st.session_state.generating = False
        st.error("Please enter a client name.")
        st.stop()
    if not st.session_state.nav_tags:
        st.session_state.generating = False
        st.error("Please add at least one tag.")
        st.stop()

    nav_categories    = st.session_state.nav_tags
    seed_items        = [x.strip() for x in seeds_input.strip().split("\n") if x.strip()]
    market_code       = MARKETS[market_label]
    tags              = ", ".join(nav_categories)
    web_note          = f" Client website: {website} — only generate content relevant to products on this site." if website else ""
    selected_lang_obj = [l for l in LANGUAGES if l["code"] in st.session_state.selected_langs]

    import base64
    genai_client = genai.Client(api_key=gemini_key)

    # ── PROMPTS MODE ─────────────────────────────────────────────
    if st.session_state.mode == "prompts":
        intent_codes = st.session_state.get("selected_intent", ["informational"])
        seed_note    = f" Must include these seed prompts: {', '.join(seed_items)}." if seed_items else ""

        INTENT_INSTRUCTIONS = {
            "informational":  "Generate INFORMATIONAL prompts — questions starting with 'What is', 'How does', 'Tell me about', 'Explain', 'Why is'. These should seek knowledge about the client's products or category.",
            "comparison":     "Generate COMPARISON prompts — questions like 'What is the best X vs Y', 'Compare X and Y', 'Which is better', 'What are the differences between'. These compare the client's products against alternatives.",
            "recommendation": "Generate RECOMMENDATION prompts — questions like 'Suggest a', 'What should I use for', 'Recommend the best', 'What is the ideal X for Y'. These seek product or service suggestions.",
            "action":         "Generate ACTION prompts — requests like 'Help me find', 'Give me a list of', 'Where can I buy', 'Show me options for'. These are task-oriented prompts seeking specific outcomes.",
        }

        def generate_prompts_for_language(lang_name, intent_code):
            prompt = f"""You are an AI search strategist. Client: "{client_name}", market: "{market_label}".{web_note}
Tags to assign (from client's website nav): {tags}.

Generate exactly {prompt_count} AI prompts in {lang_name} that users would type into ChatGPT, Gemini, Perplexity, or similar AI tools when looking for products/services like this client's.

{INTENT_INSTRUCTIONS[intent_code]}

Rules:
- Each prompt must feel natural — like something a real person would type into an AI chatbot
- Assign each prompt exactly ONE tag from: {tags}
- Assign an AI platform: one of ChatGPT, Gemini, Perplexity, Claude, or General (if platform-agnostic)
- All prompts must be in {lang_name}{seed_note}
- Add "validation": "confirmed" if clearly relevant to this client, "inferred" if loosely relevant

Return ONLY a raw JSON array, no markdown, no explanation:
[{{"prompt":"...","intent":"{intent_code}","tag":"...","platform":"ChatGPT","validation":"confirmed"}}]"""
            response = genai_client.models.generate_content(model="gemini-2.5-flash-lite", contents=prompt)
            match    = re.search(r'\[[\s\S]*\]', response.text)
            if not match:
                raise ValueError(f"Could not parse response for {lang_name} / {intent_code}")
            items = json.loads(match.group())
            for item in items:
                item["language"] = lang_name
                item["combined"] = f"{item['prompt']}, {item['intent']}, {item['tag']}"
            return items

        all_prompts  = []
        total_runs   = len(selected_lang_obj) * len(intent_codes)
        run_count    = 0
        for lang in selected_lang_obj:
            for intent_code in intent_codes:
                run_count += 1
                with st.spinner(f"⏳ Generating {lang['name']} / {intent_code} prompts ({run_count}/{total_runs})..."):
                    try:
                        all_prompts.extend(generate_prompts_for_language(lang["name"], intent_code))
                    except Exception as e:
                        st.session_state.generating = False
                        st.error(f"Error for {lang['name']} / {intent_code}: {str(e)}")
                        st.stop()

        st.session_state.generating = False
        selected_lang_names = [l["name"] for l in LANGUAGES if l["code"] in st.session_state.selected_langs]
        st.success(f"✅ {len(all_prompts)} prompts generated!")
        st.divider()

        m1, m2, m3 = st.columns(3)
        m1.metric("Total prompts", len(all_prompts))
        m2.metric("Languages", len(selected_lang_obj))
        m3.metric("Intents", ", ".join(i.title() for i in intent_codes))
        st.divider()

        import pandas as pd
        df_p = pd.DataFrame([{
            "Prompt":       p["prompt"],
            "Intent":       p["intent"],
            "Tag":          p["tag"],
            "Platform":     p.get("platform","General"),
            "Language":     p["language"],
            "Validation":   p.get("validation",""),
            "Combined Entry": p["combined"]
        } for p in all_prompts])

        st.dataframe(df_p, use_container_width=True, height=450, column_config={
            "Prompt":       st.column_config.TextColumn(width="large"),
            "Intent":       st.column_config.TextColumn(width="small"),
            "Tag":          st.column_config.TextColumn(width="medium"),
            "Platform":     st.column_config.TextColumn(width="small"),
            "Language":     st.column_config.TextColumn(width="small"),
            "Validation":   st.column_config.TextColumn(width="small"),
            "Combined Entry": st.column_config.TextColumn(width="large"),
        })
        st.divider()

        timestamp = datetime.now().strftime("%Y%m%d-%H%M")
        slug      = client_name.lower().replace(" ","-")

        # CSV
        csv_buf_p = io.StringIO()
        wp = csv.DictWriter(csv_buf_p, fieldnames=["prompt","intent","tag","platform","language","validation","combined"])
        wp.writeheader()
        for p in all_prompts:
            wp.writerow({k: p.get(k,"") for k in ["prompt","intent","tag","platform","language","validation","combined"]})

        # Excel
        wbp = Workbook(); wsp = wbp.active; wsp.title = "Prompts"
        DARK,WHITE="1A1A1A","FFFFFF"; LIGHT_GREY="F5F5F5"
        PURPLE_BG="EDE7F6"; PURPLE_TXT="4A148C"
        GREEN_BG="E8F5E9"; GREEN_TXT="1B5E20"
        AMBER_BG="FFF8E1"; AMBER_TXT="E65100"
        thin=Side(style="thin",color="E0E0E0"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

        wsp.merge_cells("A1:G1"); t=wsp["A1"]
        t.value=f"AI Prompt Research — {client_name} | {market_label} | {', '.join(i.title() for i in intent_codes)}"
        t.font=Font(name="Calibri",bold=True,size=13,color=WHITE)
        t.fill=PatternFill("solid",fgColor=DARK)
        t.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        wsp.row_dimensions[1].height=30

        wsp.merge_cells("A2:G2"); meta=wsp["A2"]
        meta.value=f"Generated {datetime.now().strftime('%d %b %Y')}  |  {', '.join(intent_codes)} intent  |  {len(all_prompts)} prompts  |  {', '.join(selected_lang_names)}"
        meta.font=Font(name="Calibri",size=10,color="888888")
        meta.fill=PatternFill("solid",fgColor="F0F0F0")
        meta.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        wsp.row_dimensions[2].height=18

        for col,h in enumerate(["Prompt","Intent","Tag","Platform","Language","Validation","Combined Entry"],1):
            c=wsp.cell(row=3,column=col,value=h)
            c.font=Font(name="Calibri",bold=True,size=10,color=WHITE)
            c.fill=PatternFill("solid",fgColor=DARK)
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=border
        wsp.row_dimensions[3].height=22

        for r,p in enumerate(all_prompts,4):
            rf=PatternFill("solid",fgColor=WHITE if r%2==0 else LIGHT_GREY)
            vals=[p.get("prompt",""),p.get("intent",""),p.get("tag",""),
                  p.get("platform",""),p.get("language",""),p.get("validation",""),p.get("combined","")]
            for col,val in enumerate(vals,1):
                c=wsp.cell(row=r,column=col,value=val)
                c.font=Font(name="Calibri",size=10)
                c.alignment=Alignment(vertical="center",wrap_text=(col in [1,7]))
                c.border=border; c.fill=rf
            # Intent colour
            ic=wsp.cell(row=r,column=2)
            ic.fill=PatternFill("solid",fgColor=PURPLE_BG)
            ic.font=Font(name="Calibri",size=10,bold=True,color=PURPLE_TXT)
            ic.alignment=Alignment(horizontal="center",vertical="center")
            # Validation colour
            vc=wsp.cell(row=r,column=6)
            if p.get("validation")=="confirmed":
                vc.fill=PatternFill("solid",fgColor=GREEN_BG); vc.font=Font(name="Calibri",size=10,color=GREEN_TXT)
            elif p.get("validation")=="inferred":
                vc.fill=PatternFill("solid",fgColor=AMBER_BG); vc.font=Font(name="Calibri",size=10,color=AMBER_TXT)
            vc.alignment=Alignment(horizontal="center",vertical="center")
            wsp.row_dimensions[r].height=22

        for i,w in enumerate([50,14,18,14,12,12,50],1):
            wsp.column_dimensions[get_column_letter(i)].width=w
        wsp.freeze_panes="A4"

        xlsx_buf_p=io.BytesIO(); wbp.save(xlsx_buf_p); xlsx_buf_p.seek(0)

        dl1,dl2=st.columns(2)
        with dl1:
            st.download_button("⬇️ Download Excel",data=xlsx_buf_p,
                file_name=f"prompts-{slug}-{market_code}-{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with dl2:
            st.download_button("⬇️ Download CSV",data=csv_buf_p.getvalue().encode("utf-8-sig"),
                file_name=f"prompts-{slug}-{market_code}-{timestamp}.csv",
                mime="text/csv",use_container_width=True)

    # ── KEYWORDS MODE ─────────────────────────────────────────────
    else:
        seed_keywords = seed_items
        seed_note     = f" Must include these seed keywords: {', '.join(seed_keywords)}." if seed_keywords else ""
        qtype_code    = st.session_state.selected_qt

    # Test DataForSEO credentials before starting
    dfs_ok = False
    if dfs_login and dfs_password:
        try:
            test_creds = base64.b64encode(f"{dfs_login}:{dfs_password}".encode()).decode()
            test_res = requests.post(
                "https://api.dataforseo.com/v3/keywords_data/google_ads/search_volume/live",
                headers={"Authorization": f"Basic {test_creds}", "Content-Type": "application/json"},
                json=[{"keywords": ["test"], "location_name": DFS_LOCATIONS.get(market_label, market_label), "language_name": "English"}],
                timeout=10
            )
            test_data = test_res.json()
            if test_data.get("status_code") == 20000:
                dfs_ok = True
            else:
                st.warning(f"⚠️ DataForSEO connected but returned error: {test_data.get('status_message', 'Unknown error')}. Keywords will generate without volumes.")
        except Exception as e:
            st.warning(f"⚠️ Could not connect to DataForSEO: {str(e)}. Keywords will generate without volumes.")

    def run_for_language(lang_name):
        prompt = f"""You are an expert SEO strategist. Client: "{client_name}", market: "{market_label}".{web_note}
Nav categories to use as tags: {tags}.
Generate exactly {branded_count} branded and {generic_count} generic keywords in {lang_name}.
{QTYPE_INSTRUCTIONS[qtype_code]}
Rules:
- Branded = includes the brand name or a clear brand variation
- Generic = category/product search with no brand name
- Assign each keyword exactly ONE tag from: {tags}
- All keywords must be in {lang_name}{seed_note}
- Add "validation": "confirmed" if clearly on-brand, "inferred" if plausible but less certain
Return ONLY a raw JSON array, no markdown, no explanation:
[{{"keyword":"...","category":"Branded","tag":"...","validation":"confirmed"}}]"""
        response = genai_client.models.generate_content(
            model="gemini-2.5-flash-lite", contents=prompt)
        match = re.search(r'\[[\s\S]*\]', response.text)
        if not match:
            raise ValueError(f"Could not parse response for {lang_name}")
        kws = json.loads(match.group())
        for kw in kws:
            kw["language"] = lang_name
            kw["volume"]   = None
            kw["combined"] = f"{kw['keyword']}, {kw['category']}, {kw['tag']}"
        return kws

    all_keywords = []
    for li, lang in enumerate(selected_lang_obj):
        with st.spinner(f"⏳ Generating {lang['name']} keywords ({li+1}/{len(selected_lang_obj)})..."):
            try:
                all_keywords.extend(run_for_language(lang["name"]))
            except Exception as e:
                st.session_state.generating = False
                st.error(f"Error for {lang['name']}: {str(e)}")
                st.stop()

    if dfs_ok:
        dfs_creds   = base64.b64encode(f"{dfs_login}:{dfs_password}".encode()).decode()
        dfs_headers = {"Authorization": f"Basic {dfs_creds}", "Content-Type": "application/json"}
        dfs_location = DFS_LOCATIONS.get(market_label, market_label)

        with st.spinner("Fetching search volumes from DataForSEO..."):
            vol_map = {}
            try:
                # Group keywords by language for accurate volume fetching
                from collections import defaultdict
                lang_groups = defaultdict(list)
                for i, kw in enumerate(all_keywords):
                    lang_groups[kw.get("language", "English")].append((i, kw["keyword"]))

                DFS_LANG_MAP = {
                    "English": "English", "Arabic": "Arabic", "French": "French",
                    "German": "German", "Hindi": "Hindi",
                    "Chinese (Simplified)": "Chinese (Simplified)", "Japanese": "Japanese"
                }

                for lang_name, kw_pairs in lang_groups.items():
                    kw_texts = [kw for _, kw in kw_pairs]
                    dfs_lang = DFS_LANG_MAP.get(lang_name, "English")
                    payload  = [{"keywords": kw_texts, "location_name": dfs_location, "language_name": dfs_lang}]
                    res  = requests.post(
                        "https://api.dataforseo.com/v3/keywords_data/google_ads/search_volume/live",
                        headers=dfs_headers, json=payload, timeout=60)
                    data  = res.json()
                    items = data.get("tasks", [{}])[0].get("result", []) or []
                    for item in items:
                        kw  = item.get("keyword")
                        vol = item.get("search_volume")
                        if kw and vol is not None:
                            vol_map[kw] = vol
            except Exception as e:
                st.warning(f"⚠️ Volume fetch failed: {str(e)}")
                vol_map = {}

            # Assign volumes — None where Google has no data
            for i, kw in enumerate(all_keywords):
                all_keywords[i]["volume"] = vol_map.get(kw["keyword"])

    st.session_state.generating = False
    st.success(f"✅ {len(all_keywords)} keywords generated!")
    st.divider()

    branded_n   = sum(1 for k in all_keywords if k["category"] == "Branded")
    generic_n   = sum(1 for k in all_keywords if k["category"] == "Generic")
    confirmed_n = sum(1 for k in all_keywords if k.get("validation") == "confirmed")
    total_vol   = sum(k.get("volume") or 0 for k in all_keywords)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total", len(all_keywords))
    m2.metric("Branded", branded_n)
    m3.metric("Generic", generic_n)
    m4.metric("Total volume" if dfs_ok else "Confirmed",
              f"{total_vol:,}" if dfs_ok else confirmed_n)
    st.divider()

    import pandas as pd
    df = pd.DataFrame([{
        "Keyword":               k["keyword"],
        "Original (if replaced)":k.get("original_keyword",""),
        "Category":              k["category"],
        "Tag":                   k["tag"],
        "Language":              k["language"],
        "Search Volume":         k.get("volume") or "",
        "Validation":            k.get("validation") or "",
        "Combined Entry":        k["combined"]
    } for k in all_keywords])

    st.dataframe(df, use_container_width=True, height=450, column_config={
        "Keyword":               st.column_config.TextColumn(width="large"),
        "Original (if replaced)":st.column_config.TextColumn(width="medium"),
        "Category":              st.column_config.TextColumn(width="small"),
        "Tag":                   st.column_config.TextColumn(width="medium"),
        "Language":              st.column_config.TextColumn(width="small"),
        "Search Volume":         st.column_config.NumberColumn(width="small"),
        "Validation":            st.column_config.TextColumn(width="small"),
        "Combined Entry":        st.column_config.TextColumn(width="large"),
    })
    st.divider()

    timestamp = datetime.now().strftime("%Y%m%d-%H%M")
    slug      = client_name.lower().replace(" ","-")

    csv_buf = io.StringIO()
    w = csv.DictWriter(csv_buf, fieldnames=["keyword","original_keyword","category","tag","language","volume","validation","combined"])
    w.writeheader()
    for kw in all_keywords:
        w.writerow({k: kw.get(k,"") for k in ["keyword","original_keyword","category","tag","language","volume","validation","combined"]})

    wb = Workbook(); ws = wb.active; ws.title = "Keywords"
    DARK,WHITE="1A1A1A","FFFFFF"; LIGHT_GREY="F5F5F5"
    BLUE_BG="E3F2FD"; BLUE_TXT="0D47A1"; GREEN_BG="E8F5E9"; GREEN_TXT="1B5E20"
    AMBER_BG="FFF8E1"; AMBER_TXT="E65100"
    thin=Side(style="thin",color="E0E0E0"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.merge_cells("A1:H1"); t=ws["A1"]
    t.value=f"Keyword Research — {client_name} | {market_label} | {', '.join(selected_lang_names)}"
    t.font=Font(name="Calibri",bold=True,size=13,color=WHITE)
    t.fill=PatternFill("solid",fgColor=DARK)
    t.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=30

    ws.merge_cells("A2:H2"); meta=ws["A2"]
    meta.value=f"Generated {datetime.now().strftime('%d %b %Y')}  |  {qtype_code}  |  {len(all_keywords)} keywords  |  {branded_n} branded / {generic_n} generic"
    meta.font=Font(name="Calibri",size=10,color="888888")
    meta.fill=PatternFill("solid",fgColor="F0F0F0")
    meta.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[2].height=18

    for col,h in enumerate(["Keyword","Original (if replaced)","Category","Tag","Language","Search Volume","Validation","Combined Entry"],1):
        c=ws.cell(row=3,column=col,value=h)
        c.font=Font(name="Calibri",bold=True,size=10,color=WHITE)
        c.fill=PatternFill("solid",fgColor=DARK)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=border
    ws.row_dimensions[3].height=22

    for r,kw in enumerate(all_keywords,4):
        rf=PatternFill("solid",fgColor=WHITE if r%2==0 else LIGHT_GREY)
        vals=[kw.get("keyword",""),kw.get("original_keyword",""),kw.get("category",""),
              kw.get("tag",""),kw.get("language",""),kw.get("volume",""),kw.get("validation",""),kw.get("combined","")]
        for col,val in enumerate(vals,1):
            c=ws.cell(row=r,column=col,value=val)
            c.font=Font(name="Calibri",size=10)
            c.alignment=Alignment(vertical="center",wrap_text=(col==8))
            c.border=border; c.fill=rf
        cat=ws.cell(row=r,column=3); is_brand=kw.get("category")=="Branded"
        cat.fill=PatternFill("solid",fgColor=BLUE_BG if is_brand else GREEN_BG)
        cat.font=Font(name="Calibri",size=10,bold=True,color=BLUE_TXT if is_brand else GREEN_TXT)
        cat.alignment=Alignment(horizontal="center",vertical="center")
        vc=ws.cell(row=r,column=7)
        if kw.get("validation")=="confirmed":
            vc.fill=PatternFill("solid",fgColor=GREEN_BG); vc.font=Font(name="Calibri",size=10,color=GREEN_TXT)
        elif kw.get("validation")=="inferred":
            vc.fill=PatternFill("solid",fgColor=AMBER_BG); vc.font=Font(name="Calibri",size=10,color=AMBER_TXT)
        vc.alignment=Alignment(horizontal="center",vertical="center")
        volc=ws.cell(row=r,column=6)
        volc.alignment=Alignment(horizontal="right",vertical="center")
        if isinstance(kw.get("volume"),int): volc.number_format="#,##0"
        ws.row_dimensions[r].height=18

    for i,w in enumerate([32,24,12,18,12,14,12,48],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A4"

    xlsx_buf=io.BytesIO(); wb.save(xlsx_buf); xlsx_buf.seek(0)

    dl1,dl2=st.columns(2)
    with dl1:
        st.download_button("⬇️ Download Excel",data=xlsx_buf,
            file_name=f"keywords-{slug}-{market_code}-{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with dl2:
        st.download_button("⬇️ Download CSV",data=csv_buf.getvalue().encode("utf-8-sig"),
            file_name=f"keywords-{slug}-{market_code}-{timestamp}.csv",
            mime="text/csv",use_container_width=True)
